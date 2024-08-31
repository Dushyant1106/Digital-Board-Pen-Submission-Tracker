import tkinter as tk
from tkinter import messagebox
from tkcalendar import Calendar
from openpyxl import load_workbook
from datetime import datetime

class PenTrackerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pen Tracker")
        self.root.configure(bg="#FFE5B4")  # Set background color to light peach
        
        self.calendar_label = tk.Label(root, text="Select Date:", bg="#FFE5B4", fg="#333333", font=("Helvetica", 12))
        self.calendar_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        
        self.cal = Calendar(root, selectmode="day", year=datetime.now().year, month=datetime.now().month, day=datetime.now().day)
        self.cal.grid(row=0, column=1, padx=10, pady=5)
        
        self.class_label = tk.Label(root, text="Class:", bg="#FFE5B4", fg="#333333", font=("Helvetica", 12))
        self.class_label.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        
        self.class_entry = tk.Entry(root)
        self.class_entry.grid(row=1, column=1, padx=10, pady=5)
        
        self.name_label = tk.Label(root, text="Representative Name:", bg="#FFE5B4", fg="#333333", font=("Helvetica", 12))
        self.name_label.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        
        self.name_entry = tk.Entry(root)
        self.name_entry.grid(row=2, column=1, padx=10, pady=5)
        
        self.submit_button = tk.Button(root, text="Verified and Submitted", command=self.submit_pen, bg="#FF6F61", fg="white", font=("Helvetica", 12, "bold"))
        self.submit_button.grid(row=3, column=0, columnspan=2, padx=10, pady=5)
        
    def submit_pen(self):
        date = self.cal.get_date()
        class_number = self.class_entry.get()
        name = self.name_entry.get()
        
        if not class_number or not name:
            messagebox.showerror("Error", "Please enter class number and representative name.")
            return
        
        # Convert class_number to integer
        try:
            class_number = int(class_number)
        except ValueError:
            messagebox.showerror("Error", "Class number must be a valid integer.")
            return
        
        try:
            wb = load_workbook("datasheet.xlsx")
            sheet = wb.active
            
            # Find the column for the selected date or append a new column if not found
            for col in range(3, sheet.max_column + 2):
                if sheet.cell(row=1, column=col).value == date:
                    date_column = col
                    break
            else:
                date_column = sheet.max_column + 1
                sheet.cell(row=1, column=date_column, value=date)
            
            # Check if the entry already exists for the given class number and representative name on the selected date
            for i in range(2, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == class_number and sheet.cell(row=i, column=2).value == name:
                    if sheet.cell(row=i, column=date_column).value == "Yes":
                        messagebox.showinfo("Already Submitted", f"{name} from class {class_number} has already submitted the pen for the selected date.")
                    else:
                        sheet.cell(row=i, column=date_column, value="Yes")
                        wb.save("datasheet.xlsx")
                        messagebox.showinfo("Success", f"{name} from class {class_number} has submitted the pen.")
                    return
            
            # If the entry doesn't exist, prompt for new class registration
            response = messagebox.askyesno("New Class Registration", "Class number and representative name not found. Do you want to register a new class?")
            if response:
                self.register_new_class(class_number, name, wb, sheet, date_column)
                
        except FileNotFoundError:
            messagebox.showerror("Error", "Excel file not found.")
            
    def register_new_class(self, class_number, name, wb, sheet, date_column):
        new_class_window = tk.Toplevel(self.root)
        new_class_window.title("Register New Class")
        new_class_window.configure(bg="#FFE5B4")  # Set background color to light peach
        
        class_label = tk.Label(new_class_window, text="Class:", bg="#FFE5B4", fg="#333333", font=("Helvetica", 12))
        class_label.grid(row=0, column=0, padx=10, pady=5)
        
        class_entry = tk.Entry(new_class_window)
        class_entry.grid(row=0, column=1, padx=10, pady=5)
        class_entry.insert(0, class_number)
        
        name_label = tk.Label(new_class_window, text="Representative Name:", bg="#FFE5B4", fg="#333333", font=("Helvetica", 12))
        name_label.grid(row=1, column=0, padx=10, pady=5)
        
        name_entry = tk.Entry(new_class_window)
        name_entry.grid(row=1, column=1, padx=10, pady=5)
        name_entry.insert(0, name)
        
        def register():
            new_class_number = class_entry.get()
            new_name = name_entry.get()
            try:
                sheet.append([int(new_class_number), new_name] + ["No"]*(date_column-3) + ["Yes"])
                messagebox.showinfo("Success", f"{new_name} from class {new_class_number} has been registered.")
                wb.save("datasheet.xlsx")  # Save the workbook after updating the data
                new_class_window.destroy()
            except FileNotFoundError:
                messagebox.showerror("Error", "Excel file not found.")
                
        register_button = tk.Button(new_class_window, text="Register", command=register, bg="#FF6F61", fg="white", font=("Helvetica", 12, "bold"))
        register_button.grid(row=2, column=0, columnspan=2, padx=10, pady=5)

root = tk.Tk()
app = PenTrackerApp(root)
root.mainloop()
