import tkinter as tk
from tkcalendar import Calendar
from openpyxl import load_workbook
from tkinter import messagebox
from datetime import datetime

class PenTrackerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Pen Submission Tracker")
        self.root.configure(bg="#FFE5B4")  # Set background color to light peach
        
        self.calendar_label = tk.Label(root, text="Select Date:", bg="#FFE5B4", fg="#333333", font=("Helvetica", 12))
        self.calendar_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        
        self.cal = Calendar(root, selectmode="day", year=datetime.now().year, month=datetime.now().month, day=datetime.now().day)
        self.cal.grid(row=0, column=1, padx=10, pady=5)
        
        self.class_label = tk.Label(root, text="Class Number:", bg="#FFE5B4", fg="#333333", font=("Helvetica", 12))
        self.class_label.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        
        self.class_entry = tk.Entry(root)
        self.class_entry.grid(row=1, column=1, padx=10, pady=5)
        
        self.name_label = tk.Label(root, text="Representative Name:", bg="#FFE5B4", fg="#333333", font=("Helvetica", 12))
        self.name_label.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        
        self.name_entry = tk.Entry(root)
        self.name_entry.grid(row=2, column=1, padx=10, pady=5)
        
        self.check_pen_button = tk.Button(root, text="Check Pen Submission", command=self.check_pen_submission, bg="#FF6F61", fg="white", font=("Helvetica", 12, "bold"))
        self.check_pen_button.grid(row=3, column=0, columnspan=2, padx=10, pady=5)
        
    def check_pen_submission(self):
        date = self.cal.get_date()
        class_number = self.class_entry.get()
        name = self.name_entry.get()
        
        if not class_number or not name:
            messagebox.showerror("Error", "Please enter class number and representative name.")
            return
        
        try:
            wb = load_workbook("datasheet.xlsx")
            sheet = wb.active
            
            # Find the column for the selected date
            date_column = None
            for col in range(3, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == date:
                    date_column = col
                    break
            
            if date_column is None:
                messagebox.showinfo("Pen Submission", f"No information found for the selected date: {date}")
                return
            
            # Search for the entry for the given class number and representative name
            pen_status = "No"
            for row in range(2, sheet.max_row + 1):
                class_num = sheet.cell(row=row, column=1).value
                rep_name = sheet.cell(row=row, column=2).value
                if class_num == int(class_number) and rep_name == name:
                    pen_status = sheet.cell(row=row, column=date_column).value
                    break
            
            if pen_status == "Yes":
                messagebox.showinfo("Pen Submission", f"The pen has been submitted by {name} from class {class_number} on {date}.")
            elif pen_status == "No":
                messagebox.showinfo("Pen Submission", f"The pen has not been submitted by {name} from class {class_number} on {date}.")
            else:
                messagebox.showinfo("Pen Submission", f"No information found for {name} from class {class_number} on {date}.")
                
        except FileNotFoundError:
            messagebox.showerror("Error", "Excel file not found.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

root = tk.Tk()
app = PenTrackerGUI(root)
root.mainloop()

